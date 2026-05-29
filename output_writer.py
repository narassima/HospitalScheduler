"""
output_writer.py
----------------
Formats and writes the output schedule Excel workbook.

Output sheets:
  • Schedule_N   — Calendar grid for each result (primary + alternates)
  • Summary_N    — Per-resource and per-department statistics
  • Violations   — Violations report (if any)
  • Index        — Overview of all schedules in the workbook
"""
import sys
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

import datetime
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from solver import InputReader, ScheduleResult, Violation


# ─────────────────────────────────────────────────────────────────────────────
# Palette helpers
# ─────────────────────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
SUBHEAD_FILL = PatternFill("solid", fgColor="2E75B6")
ALT_ROW_FILL = PatternFill("solid", fgColor="EEF4FB")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
ABSENT_FILL  = PatternFill("solid", fgColor="BFBFBF")
WEEKEND_FILL = PatternFill("solid", fgColor="F2DCDB")
UNASSIGNED_FILL = PatternFill("solid", fgColor="FCE4D6")   # orange tint
VIOLATION_FILL  = PatternFill("solid", fgColor="FFC7CE")   # red tint
OK_FILL         = PatternFill("solid", fgColor="C6EFCE")   # green tint

HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEAD_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
NORMAL_FONT  = Font(name="Calibri", size=9)
BOLD_FONT    = Font(name="Calibri", bold=True, size=9)
ITALIC_FONT  = Font(name="Calibri", italic=True, size=8, color="595959")
SMALL_FONT   = Font(name="Calibri", size=8)

THIN = Side(style="thin",   color="BFBFBF")
MED  = Side(style="medium", color="2E75B6")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED_BORDER  = Border(left=MED,  right=MED,  top=MED,  bottom=MED)

# Department colours (cycle for >10 depts)
DEPT_COLORS = [
    "4472C4", "ED7D31", "70AD47", "FFC000",
    "7030A0", "00B0F0", "FF0000", "00B050",
    "FF7F00", "9900FF", "00FFFF", "FF00FF",
]


def _dept_fill(dept_idx: int):
    hex_color = DEPT_COLORS[dept_idx % len(DEPT_COLORS)]
    return PatternFill("solid", fgColor=hex_color)


def _dept_font(dept_idx: int):
    # White text on dark colours
    dark = {"4472C4", "7030A0", "FF0000", "1F3864", "2E75B6", "9900FF"}
    color = DEPT_COLORS[dept_idx % len(DEPT_COLORS)]
    fc = "FFFFFF" if color in dark else "1F1F1F"
    return Font(name="Calibri", size=8, bold=True, color=fc)


def _cell(ws, row, col, value=None, fill=None, font=None,
          align="center", wrap=False, border=True, rotation=0):
    c = ws.cell(row=row, column=col, value=value)
    if fill:
        c.fill = fill
    c.font = font or NORMAL_FONT
    c.alignment = Alignment(horizontal=align, vertical="center",
                             wrap_text=wrap, text_rotation=rotation)
    if border:
        c.border = THIN_BORDER
    return c


def _merge_cell(ws, r1, c1, r2, c2, value, fill=None, font=None):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2,   end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    c.fill = fill or HEADER_FILL
    c.font = font or HEADER_FONT
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    c.border = MED_BORDER
    return c


# ─────────────────────────────────────────────────────────────────────────────
# Writer
# ─────────────────────────────────────────────────────────────────────────────

class OutputWriter:
    def __init__(self, reader: InputReader, results: List[ScheduleResult]):
        self.reader  = reader
        self.results = results
        self.wb      = Workbook()
        self.wb.remove(self.wb.active)     # remove default blank sheet
        self._dept_idx: Dict[str, int] = {
            dept.name: i
            for i, dept in enumerate(reader.departments)
        }

    def write(self, output_path):
        # output_path may be a file-path string or a BytesIO buffer
        # Index sheet first
        self._build_index_sheet()

        # Violations sheet (if any result has violations)
        all_violations = []
        for res in self.results:
            all_violations.extend(res.violations)
        if all_violations:
            self._build_violations_sheet(all_violations)

        # Schedule + Summary for each result
        for result in self.results:
            label = (f"Schedule" if result.alternate_index == 0
                     else f"Alt_{result.alternate_index}")
            self._build_schedule_sheet(result, label)
            self._build_summary_sheet(result, label)

        self.wb.save(output_path)
        if isinstance(output_path, str):
            print(f"Output saved -> {output_path}")

    # ── Index ────────────────────────────────────────────────────────────────
    def _build_index_sheet(self):
        ws = self.wb.create_sheet("📋 Index")
        ws.sheet_properties.tabColor = "1F3864"
        ws.sheet_view.showGridLines   = False
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 40

        _merge_cell(ws, 1, 1, 1, 4,
                    "📋  TIMETABLE SCHEDULER — OUTPUT WORKBOOK INDEX")
        ws.row_dimensions[1].height = 32

        for c, h in enumerate(
                ["Sheet", "Status", "Violations", "Notes"], 1):
            _cell(ws, 2, c, h, fill=SUBHEAD_FILL, font=SUBHEAD_FONT)
        ws.row_dimensions[2].height = 20

        r = 3
        for result in self.results:
            label = (f"Schedule" if result.alternate_index == 0
                     else f"Alt_{result.alternate_index}")
            status = "✅ Fully Feasible" if result.is_feasible else "⚠ Violations Present"
            nv = len(result.violations)
            note = (f"Primary schedule" if result.alternate_index == 0
                    else f"Alternate schedule #{result.alternate_index}")
            fill = OK_FILL if result.is_feasible else VIOLATION_FILL
            _cell(ws, r, 1, label,  fill=fill, font=BOLD_FONT, align="left")
            _cell(ws, r, 2, status, fill=fill, font=NORMAL_FONT, align="left")
            _cell(ws, r, 3, nv,     fill=fill, font=NORMAL_FONT)
            _cell(ws, r, 4, note,   fill=fill, font=NORMAL_FONT, align="left")
            ws.row_dimensions[r].height = 18
            r += 1

        # Legend
        r += 1
        _merge_cell(ws, r, 1, r, 4, "DEPARTMENT COLOUR LEGEND",
                    fill=SUBHEAD_FILL)
        ws.row_dimensions[r].height = 20
        r += 1
        for dept in self.reader.departments:
            idx  = self._dept_idx[dept.name]
            fill = _dept_fill(idx)
            font = _dept_font(idx)
            _cell(ws, r, 1, dept.name, fill=fill, font=font,
                  align="left")
            ws.merge_cells(start_row=r, start_column=1,
                           end_row=r, end_column=4)
            ws.row_dimensions[r].height = 16
            r += 1

    # ── Violations ───────────────────────────────────────────────────────────
    def _build_violations_sheet(self, violations: List[Violation]):
        ws = self.wb.create_sheet("⚠ Violations")
        ws.sheet_properties.tabColor = "C00000"
        ws.sheet_view.showGridLines   = False

        col_widths = [10, 22, 22, 22, 14, 55, 12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        _merge_cell(ws, 1, 1, 1, 7,
                    "⚠  VIOLATIONS REPORT — Constraints that could not be fully satisfied")
        ws.row_dimensions[1].height = 28

        headers = ["ID", "Type", "Resource", "Department",
                   "Date", "Details", "Severity"]
        for c, h in enumerate(headers, 1):
            _cell(ws, 2, c, h, fill=SUBHEAD_FILL, font=SUBHEAD_FONT)
        ws.row_dimensions[2].height = 20

        sev_fills = {"High": PatternFill("solid", fgColor="FFC7CE"),
                     "Medium": PatternFill("solid", fgColor="FFEB9C"),
                     "Low":  PatternFill("solid", fgColor="C6EFCE")}

        for i, v in enumerate(violations):
            r    = i + 3
            fill = sev_fills.get(v.severity, WHITE_FILL)
            _cell(ws, r, 1, v.violation_id, fill=fill, font=BOLD_FONT)
            _cell(ws, r, 2, v.vtype,        fill=fill, font=NORMAL_FONT, align="left")
            _cell(ws, r, 3, v.resource,     fill=fill, font=NORMAL_FONT, align="left")
            _cell(ws, r, 4, v.department,   fill=fill, font=NORMAL_FONT, align="left")
            _cell(ws, r, 5, v.date,         fill=fill, font=NORMAL_FONT)
            _cell(ws, r, 6, v.details,      fill=fill, font=NORMAL_FONT,
                  align="left", wrap=True)
            _cell(ws, r, 7, v.severity,     fill=fill, font=BOLD_FONT)
            ws.row_dimensions[r].height = 20

    # ── Schedule calendar ────────────────────────────────────────────────────
    def _build_schedule_sheet(self, result: ScheduleResult, label: str):
        sheet_name = f"📅 {label}"
        ws = self.wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = (
            "2E75B6" if result.alternate_index == 0 else "70AD47")
        ws.sheet_view.showGridLines = False

        resources = self.reader.resources
        dates     = self.reader.dates

        # Fixed columns
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 22

        # Date columns
        for j in range(len(dates)):
            ws.column_dimensions[get_column_letter(j + 3)].width = 5

        status_str = ("✅ Fully Feasible" if result.is_feasible
                      else f"⚠ {len(result.violations)} Violation(s)")
        _merge_cell(ws, 1, 1, 1, 2 + len(dates),
                    f"📅  {label.upper()} — {status_str}")
        ws.row_dimensions[1].height = 28

        # Month / week grouping sub-header
        _cell(ws, 2, 1, "ID",   fill=SUBHEAD_FILL, font=SUBHEAD_FONT)
        _cell(ws, 2, 2, "Name", fill=SUBHEAD_FILL, font=SUBHEAD_FONT)
        for j, dt in enumerate(dates):
            col = j + 3
            cell = _cell(ws, 2, col, dt,
                         fill=SUBHEAD_FILL, font=SUBHEAD_FONT,
                         rotation=90)
            cell.number_format = "DD-MMM"
        ws.row_dimensions[2].height = 64

        # Day-of-week sub-row
        dow_fills = {5: WEEKEND_FILL, 6: WEEKEND_FILL}
        dow_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        for j, dt in enumerate(dates):
            col  = j + 3
            fill = dow_fills.get(dt.weekday(), ALT_ROW_FILL)
            _cell(ws, 3, col, dow_names[dt.weekday()],
                  fill=fill,
                  font=Font(name="Calibri", size=7, bold=(dt.weekday() >= 5)),
                  rotation=90)
        _cell(ws, 3, 1, "",   fill=ALT_ROW_FILL)
        _cell(ws, 3, 2, "",   fill=ALT_ROW_FILL)
        ws.row_dimensions[3].height = 28

        # Resource rows
        for i, res in enumerate(resources):
            r    = i + 4
            fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
            _cell(ws, r, 1, res.id,   fill=fill, font=BOLD_FONT)
            _cell(ws, r, 2, res.name, fill=fill, font=BOLD_FONT, align="left")

            for j, dt in enumerate(dates):
                col     = j + 3
                absent  = self.reader.absence.get((res.name, dt), False)
                weekend = dt.weekday() >= 5

                if absent:
                    cell = _cell(ws, r, col, "A",
                                 fill=ABSENT_FILL if not weekend else WEEKEND_FILL,
                                 font=Font(name="Calibri", size=7,
                                           color="595959"))
                else:
                    # Find assigned dept
                    assigned = None
                    for dept in self.reader.departments:
                        if result.assignment.get((res.name, dept.name, dt), 0) == 1:
                            assigned = dept
                            break
                    if assigned:
                        idx  = self._dept_idx[assigned.name]
                        abbr = assigned.name[:3].upper()
                        cell = _cell(ws, r, col, abbr,
                                     fill=_dept_fill(idx),
                                     font=_dept_font(idx))
                    else:
                        # Available but unassigned
                        cell = _cell(ws, r, col, "—",
                                     fill=UNASSIGNED_FILL,
                                     font=Font(name="Calibri", size=7,
                                               color="BFBFBF"))

            ws.row_dimensions[r].height = 16

        # Footer legend
        leg_r = len(resources) + 4
        ws.cell(row=leg_r, column=1,
                value=("A = Absent  |  Colour = Dept assigned  |  "
                       "— = Available but unassigned  |  "
                       + "  ".join(
                            f"{dept.name[:3].upper()}={dept.name}"
                            for dept in self.reader.departments)
                       )).font = ITALIC_FONT

    # ── Summary ──────────────────────────────────────────────────────────────
    def _build_summary_sheet(self, result: ScheduleResult, label: str):
        sheet_name = f"📊 Summary_{label}"
        ws = self.wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = "ED7D31"
        ws.sheet_view.showGridLines  = False

        resources = self.reader.resources
        depts     = self.reader.departments
        dates     = self.reader.dates

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 14
        for i, dept in enumerate(depts):
            ws.column_dimensions[get_column_letter(6 + i)].width = 14

        _merge_cell(ws, 1, 1, 1, 5 + len(depts),
                    f"📊  SUMMARY — {label.upper()}")
        ws.row_dimensions[1].height = 28

        # Header
        base_cols = ["ID", "Name", "Total\nWork Days",
                     "Total\nWork Hours", "Absent\nDays"]
        for c, h in enumerate(base_cols, 1):
            _cell(ws, 2, c, h, fill=SUBHEAD_FILL, font=SUBHEAD_FONT, wrap=True)
        for i, dept in enumerate(depts):
            _cell(ws, 2, 6 + i, f"Days in\n{dept.name}",
                  fill=_dept_fill(self._dept_idx[dept.name]),
                  font=_dept_font(self._dept_idx[dept.name]), wrap=True)
        ws.row_dimensions[2].height = 36

        # Per-resource rows
        for i, res in enumerate(resources):
            r    = i + 3
            fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL

            total_days  = sum(
                result.assignment.get((res.name, dept.name, dt), 0)
                for dept in depts for dt in dates)
            total_hours = total_days * res.daily_hours
            absent_days = sum(
                1 for dt in dates
                if self.reader.absence.get((res.name, dt), False))

            _cell(ws, r, 1, res.id,          fill=fill, font=BOLD_FONT)
            _cell(ws, r, 2, res.name,         fill=fill, font=NORMAL_FONT, align="left")
            _cell(ws, r, 3, total_days,       fill=fill, font=NORMAL_FONT)
            h_fill = (OK_FILL if res.min_hours <= total_hours <= res.max_hours
                      else VIOLATION_FILL)
            _cell(ws, r, 4, total_hours,      fill=h_fill, font=BOLD_FONT)
            _cell(ws, r, 5, absent_days,      fill=fill, font=NORMAL_FONT)

            for j, dept in enumerate(depts):
                days_in = sum(
                    result.assignment.get((res.name, dept.name, dt), 0)
                    for dt in dates)
                mn, mx = self.reader.dept_minmax.get(
                    (res.name, dept.name), (0, 9999))
                d_fill = (VIOLATION_FILL if (days_in < mn or days_in > mx)
                          else fill)
                _cell(ws, r, 6 + j, days_in, fill=d_fill, font=NORMAL_FONT)

            ws.row_dimensions[r].height = 16

        # Dept totals section
        dept_r = len(resources) + 4
        _merge_cell(ws, dept_r, 1, dept_r, 5 + len(depts),
                    "DEPARTMENT TOTALS",
                    fill=SUBHEAD_FILL)
        ws.row_dimensions[dept_r].height = 20
        dept_r += 1

        for c, h in enumerate(
                ["ID", "Department", "Total Hours", "Min Required",
                 "Max Allowed", "Status"], 1):
            _cell(ws, dept_r, c, h, fill=SUBHEAD_FILL, font=SUBHEAD_FONT)
        ws.row_dimensions[dept_r].height = 18
        dept_r += 1

        for dept in depts:
            total_hours = sum(
                result.assignment.get((res.name, dept.name, dt), 0)
                * res.daily_hours
                for res in resources for dt in dates)
            ok = dept.min_hours <= total_hours <= dept.max_hours
            fill = OK_FILL if ok else VIOLATION_FILL
            _cell(ws, dept_r, 1, dept.id,         fill=fill, font=BOLD_FONT)
            _cell(ws, dept_r, 2, dept.name,        fill=fill, font=NORMAL_FONT, align="left")
            _cell(ws, dept_r, 3, total_hours,      fill=fill, font=BOLD_FONT)
            _cell(ws, dept_r, 4, dept.min_hours,   fill=WHITE_FILL, font=NORMAL_FONT)
            _cell(ws, dept_r, 5, dept.max_hours,   fill=WHITE_FILL, font=NORMAL_FONT)
            status = "✅ OK" if ok else "❌ Violation"
            _cell(ws, dept_r, 6, status,           fill=fill, font=BOLD_FONT)
            ws.row_dimensions[dept_r].height = 16
            dept_r += 1
