"""
create_sample_input.py
----------------------
Creates a realistic pre-filled sample input Excel file for demonstration.

Scenario: 5 doctors, 3 departments (Cardiology, Neurology, Orthopedics),
          plan period = June 2026.

Run:
    python create_sample_input.py
"""
import sys
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

import datetime
import sys
import os

# Add the project directory to the path
sys.path.insert(0, os.path.dirname(__file__))

from generate_template import generate_template
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

WEEKEND_FILL = PatternFill("solid", fgColor="F2DCDB")
ABSENT_FILL  = PatternFill("solid", fgColor="BFBFBF")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
ALT_FILL     = PatternFill("solid", fgColor="EEF4FB")

THIN = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def create_sample():
    OUTPUT = "sample_input.xlsx"
    START  = datetime.date(2026, 6, 1)
    END    = datetime.date(2026, 6, 30)
    N_RES  = 5
    N_DEPT = 3

    # First generate the blank template
    generate_template(OUTPUT, N_RES, N_DEPT, START, END)

    # Now open it and fill in realistic sample data
    wb = openpyxl.load_workbook(OUTPUT)

    # ── Config sheet ──────────────────────────────────────────────────────────
    ws = wb["Config"]
    ws.cell(row=3, column=2).value = 5
    ws.cell(row=4, column=2).value = START
    ws.cell(row=5, column=2).value = END
    ws.cell(row=6, column=2).value = 3
    ws.cell(row=7, column=2).value = 2

    # ── Resources sheet ───────────────────────────────────────────────────────
    ws = wb["Resources"]
    doctors = [
        ("Dr. Sarah Chen",    64, 160, 8),
        ("Dr. James Patel",   64, 160, 8),
        ("Dr. Maria Santos",  40, 120, 8),   # part-time
        ("Dr. Raj Kumar",     64, 160, 8),
        ("Dr. Emily Nguyen",  80, 176, 8),
    ]
    for i, (name, minh, maxh, daily) in enumerate(doctors):
        r = i + 3
        ws.cell(row=r, column=2).value = name
        ws.cell(row=r, column=3).value = minh
        ws.cell(row=r, column=4).value = maxh
        ws.cell(row=r, column=5).value = daily

    # ── Departments sheet ─────────────────────────────────────────────────────
    ws = wb["Departments"]
    depts = [
        ("Cardiology",  160, 400),
        ("Neurology",   80,  300),
        ("Orthopedics", 80,  300),
    ]
    for i, (name, minh, maxh) in enumerate(depts):
        r = i + 3
        ws.cell(row=r, column=2).value = name
        ws.cell(row=r, column=3).value = minh
        ws.cell(row=r, column=4).value = maxh

    # ── Absence sheet — add some extra absences ───────────────────────────────
    ws = wb["Absence"]
    # Dr. Sarah Chen takes vacation June 16-20
    # Dr. James Patel is absent June 10 (conference)
    dates = [START + datetime.timedelta(days=i)
             for i in range((END - START).days + 1)]

    # Find the date columns: col 3 onwards = dates
    # Row 3 = first resource (Dr. Sarah Chen), row 4 = Dr. James, etc.
    doctor_names = [d[0] for d in doctors]

    for row in range(3, 3 + N_RES):
        rname = ws.cell(row=row, column=2).value
        for col_idx, dt in enumerate(dates):
            col = col_idx + 3
            cell = ws.cell(row=row, column=col)
            # Check if already absent (weekend)
            if cell.value == "Absent":
                continue
            # Add specific absences
            if rname == "Dr. Sarah Chen" and datetime.date(2026, 6, 16) <= dt <= datetime.date(2026, 6, 19):
                cell.value = "Absent"
                cell.fill  = ABSENT_FILL
                cell.font  = Font(name="Calibri", color="595959", bold=True, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           text_rotation=90)
                cell.border = THIN_BORDER
            elif rname == "Dr. James Patel" and dt == datetime.date(2026, 6, 10):
                cell.value = "Absent"
                cell.fill  = ABSENT_FILL
                cell.font  = Font(name="Calibri", color="595959", bold=True, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           text_rotation=90)
                cell.border = THIN_BORDER
            elif rname == "Dr. Maria Santos" and dt == datetime.date(2026, 6, 25):
                cell.value = "Absent"
                cell.fill  = ABSENT_FILL
                cell.font  = Font(name="Calibri", color="595959", bold=True, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           text_rotation=90)
                cell.border = THIN_BORDER

    # ── Dept_Mapping sheet ────────────────────────────────────────────────────
    ws = wb["Dept_Mapping"]
    # (resource_row, dept_col): 1 = can, 0 = cannot
    # Dept cols: 3=Cardiology, 4=Neurology, 5=Orthopedics
    # Dr. Sarah Chen — all depts
    # Dr. James Patel — Cardiology + Neurology
    # Dr. Maria Santos — Neurology + Orthopedics (no Cardiology)
    # Dr. Raj Kumar — all depts
    # Dr. Emily Nguyen — Cardiology + Orthopedics
    mappings = [
        [1, 1, 1],   # Sarah
        [1, 1, 0],   # James
        [0, 1, 1],   # Maria
        [1, 1, 1],   # Raj
        [1, 0, 1],   # Emily
    ]
    for i, mapping in enumerate(mappings):
        r = i + 3
        ws.cell(row=r, column=3).value = mapping[0]
        ws.cell(row=r, column=4).value = mapping[1]
        ws.cell(row=r, column=5).value = mapping[2]

    # ── Dept_MinMax sheet ─────────────────────────────────────────────────────
    ws = wb["Dept_MinMax"]
    # Format: resource i, dept j → min/max days
    # The sheet has N_RES * N_DEPT rows starting at row 3
    dept_names = [d[0] for d in depts]
    minmax = {
        ("Dr. Sarah Chen",   "Cardiology"):   (4,  15),
        ("Dr. Sarah Chen",   "Neurology"):    (2,  10),
        ("Dr. Sarah Chen",   "Orthopedics"):  (2,  10),
        ("Dr. James Patel",  "Cardiology"):   (5,  15),
        ("Dr. James Patel",  "Neurology"):    (3,  10),
        ("Dr. James Patel",  "Orthopedics"):  (0,   0),   # can't do Ortho
        ("Dr. Maria Santos", "Cardiology"):   (0,   0),   # can't do Cardio
        ("Dr. Maria Santos", "Neurology"):    (3,  10),
        ("Dr. Maria Santos", "Orthopedics"):  (2,  10),
        ("Dr. Raj Kumar",    "Cardiology"):   (5,  15),
        ("Dr. Raj Kumar",    "Neurology"):    (2,  10),
        ("Dr. Raj Kumar",    "Orthopedics"):  (2,  10),
        ("Dr. Emily Nguyen", "Cardiology"):   (6,  15),
        ("Dr. Emily Nguyen", "Neurology"):    (0,   0),   # can't do Neuro
        ("Dr. Emily Nguyen", "Orthopedics"):  (3,  10),
    }

    r = 3
    for dname in doctor_names:
        for dept in dept_names:
            ws.cell(row=r, column=2).value = dname
            ws.cell(row=r, column=3).value = dept
            mn, mx = minmax.get((dname, dept), (0, 999))
            ws.cell(row=r, column=4).value = mn
            ws.cell(row=r, column=5).value = mx
            r += 1

    wb.save(OUTPUT)
    print(f"✅  Sample input saved → {OUTPUT}")
    print(f"    Scenario  : 5 doctors, 3 departments, June 2026")
    print(f"    Run scheduler with:")
    print(f"      python scheduler.py --input {OUTPUT} --output sample_output.xlsx")


if __name__ == "__main__":
    create_sample()
