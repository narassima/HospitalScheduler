"""
generate_template.py
--------------------
Generates a blank (but pre-configured) Excel input template for the
Timetable Scheduler.  Run this script first, fill in the workbook, then
run scheduler.py to produce the output schedule.

Usage:
    python generate_template.py [--output INPUT_TEMPLATE.xlsx]
"""
import sys
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

import argparse
import datetime
import sys
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.differential import DifferentialStyle


# ─────────────────────────────────────────────────────────────────────────────
# Colour palette
# ─────────────────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
SUBHEAD_FILL  = PatternFill("solid", fgColor="2E75B6")   # mid blue
ACCENT_FILL   = PatternFill("solid", fgColor="D6E4F7")   # light blue
ABSENT_FILL   = PatternFill("solid", fgColor="BFBFBF")   # grey
WEEKEND_FILL  = PatternFill("solid", fgColor="F2DCDB")   # light red
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
ALT_ROW_FILL  = PatternFill("solid", fgColor="EEF4FB")   # very light blue

HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEAD_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
NORMAL_FONT  = Font(name="Calibri", size=10)
BOLD_FONT    = Font(name="Calibri", bold=True, size=10)
ITALIC_FONT  = Font(name="Calibri", italic=True, size=9, color="595959")

THIN = Side(style="thin", color="BFBFBF")
MED  = Side(style="medium", color="2E75B6")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED_BORDER  = Border(left=MED,  right=MED,  top=MED,  bottom=MED)


def _hdr(ws, row, col, value, fill=None, font=None, align="center", wrap=False):
    """Write a styled header cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill   = fill  or HEADER_FILL
    cell.font   = font  or HEADER_FONT
    cell.alignment = Alignment(horizontal=align, vertical="center",
                               wrap_text=wrap)
    cell.border = THIN_BORDER
    return cell


def _data(ws, row, col, value=None, fill=None, font=None,
          align="left", wrap=False, border=True):
    """Write a styled data cell."""
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    cell.font = font or NORMAL_FONT
    cell.alignment = Alignment(horizontal=align, vertical="center",
                               wrap_text=wrap)
    if border:
        cell.border = THIN_BORDER
    return cell


def _merge_hdr(ws, r1, c1, r2, c2, value, fill=None, font=None):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2,   end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    cell.fill = fill or HEADER_FILL
    cell.font = font or HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=True)
    cell.border = MED_BORDER
    return cell


# ─────────────────────────────────────────────────────────────────────────────
# Sheet builders
# ─────────────────────────────────────────────────────────────────────────────

def build_config_sheet(ws):
    """Sheet 1 — Config"""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 40

    _merge_hdr(ws, 1, 1, 1, 3,
               "⚙  TIMETABLE SCHEDULER — CONFIGURATION",
               fill=PatternFill("solid", fgColor="1F3864"))

    ws.row_dimensions[1].height = 32

    headers = ["Parameter", "Value", "Notes / Instructions"]
    for c, h in enumerate(headers, 1):
        _hdr(ws, 2, c, h, fill=SUBHEAD_FILL)
    ws.row_dimensions[2].height = 20

    rows = [
        ("Number of Resources",
         10,
         "Total people to schedule (e.g. doctors)"),
        ("Plan Start Date",
         datetime.date.today().replace(day=1),
         "Format: YYYY-MM-DD  (e.g. 2026-06-01)"),
        ("Plan End Date",
         (datetime.date.today().replace(day=1) +
          datetime.timedelta(days=31)).replace(day=1) - datetime.timedelta(days=1),
         "Format: YYYY-MM-DD  (e.g. 2026-06-30)"),
        ("Number of Departments",
         4,
         "Total departments/courses/tasks"),
        ("Max Alternate Schedules to Generate",
         3,
         "Maximum number of alternative valid schedules (0 = only best)"),
    ]

    for i, (param, val, note) in enumerate(rows):
        r = i + 3
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        _data(ws, r, 1, param, fill=fill, font=BOLD_FONT)
        cell = _data(ws, r, 2, val, fill=PatternFill("solid", fgColor="FFFACD"),
                     align="center")
        if isinstance(val, datetime.date):
            cell.number_format = "YYYY-MM-DD"
        _data(ws, r, 3, note, fill=fill, font=ITALIC_FONT)
        ws.row_dimensions[r].height = 18

    ws.row_dimensions[8].height = 28
    ws.cell(row=8, column=1,
            value="⚠  After filling this sheet, run generate_template.py "
                  "--refresh to rebuild the other sheets with correct "
                  "row/column counts.").font = Font(
        name="Calibri", italic=True, color="C00000", size=9)


def build_resources_sheet(ws, n_resources=10):
    """Sheet 2 — Resources"""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    _merge_hdr(ws, 1, 1, 1, 5, "👤  RESOURCES / PERSONNEL")
    ws.row_dimensions[1].height = 28

    cols = ["ID", "Resource Name", "Min Work Hours\n(Plan Period)",
            "Max Work Hours\n(Plan Period)", "Daily Work Hours"]
    for c, h in enumerate(cols, 1):
        _hdr(ws, 2, c, h, fill=SUBHEAD_FILL, wrap=True)
    ws.row_dimensions[2].height = 30

    for i in range(1, n_resources + 1):
        r = i + 2
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        _data(ws, r, 1, i,       fill=fill, align="center", font=BOLD_FONT)
        _data(ws, r, 2, f"Resource {i}", fill=fill)
        _data(ws, r, 3, 80,      fill=fill, align="center")
        _data(ws, r, 4, 176,     fill=fill, align="center")
        _data(ws, r, 5, 8,       fill=fill, align="center")
        ws.row_dimensions[r].height = 16

    note_row = n_resources + 3
    ws.cell(row=note_row, column=1,
            value="ℹ  Replace 'Resource N' with actual names. "
                  "Hours refer to total hours in the plan period.").font = ITALIC_FONT


def build_departments_sheet(ws, n_depts=4):
    """Sheet 3 — Departments"""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    _merge_hdr(ws, 1, 1, 1, 4, "🏥  DEPARTMENTS / COURSES / TASKS")
    ws.row_dimensions[1].height = 28

    cols = ["ID", "Department Name",
            "Min Total Hours\n(Plan Period)", "Max Total Hours\n(Plan Period)"]
    for c, h in enumerate(cols, 1):
        _hdr(ws, 2, c, h, fill=SUBHEAD_FILL, wrap=True)
    ws.row_dimensions[2].height = 30

    dept_examples = ["Department A", "Department B", "Department C", "Department D"]
    for i in range(1, n_depts + 1):
        r = i + 2
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        name = dept_examples[i - 1] if i <= len(dept_examples) else f"Department {i}"
        _data(ws, r, 1, i,    fill=fill, align="center", font=BOLD_FONT)
        _data(ws, r, 2, name, fill=fill)
        _data(ws, r, 3, 100,  fill=fill, align="center")
        _data(ws, r, 4, 500,  fill=fill, align="center")
        ws.row_dimensions[r].height = 16


def build_absence_sheet(ws, n_resources=10,
                        start_date: datetime.date = None,
                        end_date:   datetime.date = None):
    """Sheet 4 — Absence calendar."""
    if start_date is None:
        start_date = datetime.date.today().replace(day=1)
    if end_date is None:
        end_date = (start_date.replace(day=1) +
                    datetime.timedelta(days=31)).replace(day=1) - datetime.timedelta(days=1)

    dates = []
    d = start_date
    while d <= end_date:
        dates.append(d)
        d += datetime.timedelta(days=1)

    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 24

    _merge_hdr(ws, 1, 1, 1, 2 + len(dates),
               "📅  ABSENCE CALENDAR  —  Mark additional absent days as 'Absent'. "
               "Saturdays & Sundays are pre-filled.")
    ws.row_dimensions[1].height = 28

    # Sub-header: resource cols
    _hdr(ws, 2, 1, "ID",   fill=SUBHEAD_FILL)
    _hdr(ws, 2, 2, "Name", fill=SUBHEAD_FILL)

    # Date headers
    for j, dt in enumerate(dates):
        col = j + 3
        cell = _hdr(ws, 2, col, dt, fill=SUBHEAD_FILL)
        cell.number_format = "DD-MMM"
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   text_rotation=90)
        ws.column_dimensions[get_column_letter(col)].width = 5

    ws.row_dimensions[2].height = 64

    # Data validation for absence cells
    dv = DataValidation(type="list", formula1='"Absent,Available"',
                        allow_blank=True, showDropDown=False)
    dv.showErrorMessage = True
    dv.errorTitle = "Invalid entry"
    dv.error = "Please select 'Absent' or leave blank (= Available)"
    ws.add_data_validation(dv)

    for i in range(1, n_resources + 1):
        r = i + 2
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        _data(ws, r, 1, i, fill=fill, align="center", font=BOLD_FONT)
        _data(ws, r, 2, f"Resource {i}", fill=fill)

        for j, dt in enumerate(dates):
            col = j + 3
            is_weekend = dt.weekday() >= 5  # Sat=5, Sun=6
            if is_weekend:
                cell = ws.cell(row=r, column=col, value="Absent")
                cell.fill   = WEEKEND_FILL
                cell.font   = Font(name="Calibri", color="C00000",
                                   bold=True, size=9)
            else:
                cell = ws.cell(row=r, column=col, value=None)
                cell.fill = fill
                cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       text_rotation=90)
            cell.border = THIN_BORDER
            col_letter = get_column_letter(col)
            dv.add(f"{col_letter}{r}")

        ws.row_dimensions[r].height = 16

    # Legend
    note_row = n_resources + 3
    ws.cell(row=note_row, column=1,
            value="Legend:  Pink = Weekend (pre-filled Absent). "
                  "Select 'Absent' from dropdown for other absent days. "
                  "Blank = Available.").font = ITALIC_FONT


def build_mapping_sheet(ws, n_resources=10, n_depts=4):
    """Sheet 5 — Dept_Mapping (feasibility matrix)."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 24

    _merge_hdr(ws, 1, 1, 1, 2 + n_depts,
               "🔗  RESOURCE–DEPARTMENT MAPPING  "
               "  (1 = Can be assigned, 0 = Cannot be assigned)")
    ws.row_dimensions[1].height = 28

    _hdr(ws, 2, 1, "ID",   fill=SUBHEAD_FILL)
    _hdr(ws, 2, 2, "Name", fill=SUBHEAD_FILL)

    for d in range(1, n_depts + 1):
        _hdr(ws, 2, d + 2, f"Department {d}", fill=SUBHEAD_FILL)
        ws.column_dimensions[get_column_letter(d + 2)].width = 16

    ws.row_dimensions[2].height = 20

    # Data validation: 0 or 1
    dv = DataValidation(type="whole", operator="between",
                        formula1="0", formula2="1",
                        allow_blank=False)
    dv.showErrorMessage = True
    dv.errorTitle = "Invalid value"
    dv.error = "Enter 1 (can assign) or 0 (cannot assign)"
    ws.add_data_validation(dv)

    for i in range(1, n_resources + 1):
        r = i + 2
        fill = ALT_ROW_FILL if i % 2 == 0 else WHITE_FILL
        _data(ws, r, 1, i, fill=fill, align="center", font=BOLD_FONT)
        _data(ws, r, 2, f"Resource {i}", fill=fill)
        for d in range(1, n_depts + 1):
            cell = _data(ws, r, d + 2, 1, fill=fill, align="center")
            dv.add(cell)
        ws.row_dimensions[r].height = 16


def build_dept_minmax_sheet(ws, n_resources=10, n_depts=4):
    """Sheet 6 — Dept_MinMax (per-resource per-dept day constraints)."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    _merge_hdr(ws, 1, 1, 1, 5,
               "📊  MIN / MAX DAYS PER RESOURCE PER DEPARTMENT  "
               "(Leave blank = no constraint)")
    ws.row_dimensions[1].height = 28

    cols = ["Resource ID", "Resource Name", "Department Name",
            "Min Days", "Max Days"]
    for c, h in enumerate(cols, 1):
        _hdr(ws, 2, c, h, fill=SUBHEAD_FILL)
    ws.row_dimensions[2].height = 20

    r = 3
    for i in range(1, n_resources + 1):
        for d in range(1, n_depts + 1):
            fill = ALT_ROW_FILL if r % 2 == 0 else WHITE_FILL
            _data(ws, r, 1, i,               fill=fill, align="center", font=BOLD_FONT)
            _data(ws, r, 2, f"Resource {i}",    fill=fill)
            _data(ws, r, 3, f"Department {d}",  fill=fill)
            _data(ws, r, 4, 0,               fill=fill, align="center")
            _data(ws, r, 5, 999,             fill=fill, align="center")
            ws.row_dimensions[r].height = 15
            r += 1

    ws.cell(row=r + 1, column=1,
            value="ℹ  Min Days = minimum days this resource MUST work in this dept. "
                  "Max Days = maximum allowed.  "
                  "Set Min=0, Max=999 to apply no constraint.").font = ITALIC_FONT


def build_working_hours_note(ws):
    """Sheet 7 — Reminder that daily hours are set in Resources sheet."""
    ws.sheet_view.showGridLines = False
    _merge_hdr(ws, 1, 1, 1, 3,
               "⏱  DAILY WORKING HOURS — see the 'Resources' sheet column E")
    ws.row_dimensions[1].height = 28
    ws.cell(row=3, column=1,
            value="Daily working hours per resource are configured in the "
                  "'Resources' sheet, column E (Daily Work Hours).  "
                  "This sheet is reserved for future per-day overrides.").font = ITALIC_FONT
    ws.column_dimensions["A"].width = 80


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def generate_template(output_path: str,
                      n_resources: int = 10,
                      n_depts: int = 4,
                      start_date: datetime.date = None,
                      end_date: datetime.date = None):

    if start_date is None:
        start_date = datetime.date.today().replace(day=1)
    if end_date is None:
        end_date = (start_date.replace(day=1) +
                    datetime.timedelta(days=31)).replace(day=1) - datetime.timedelta(days=1)

    wb = Workbook()

    # ── Sheet 1: Config ──────────────────────────────────────────────────────
    ws_cfg = wb.active
    ws_cfg.title = "Config"
    build_config_sheet(ws_cfg)

    # ── Sheet 2: Resources ───────────────────────────────────────────────────
    ws_res = wb.create_sheet("Resources")
    build_resources_sheet(ws_res, n_resources)

    # ── Sheet 3: Departments ─────────────────────────────────────────────────
    ws_dpt = wb.create_sheet("Departments")
    build_departments_sheet(ws_dpt, n_depts)

    # ── Sheet 4: Absence ─────────────────────────────────────────────────────
    ws_abs = wb.create_sheet("Absence")
    build_absence_sheet(ws_abs, n_resources, start_date, end_date)

    # ── Sheet 5: Dept_Mapping ────────────────────────────────────────────────
    ws_map = wb.create_sheet("Dept_Mapping")
    build_mapping_sheet(ws_map, n_resources, n_depts)

    # ── Sheet 6: Dept_MinMax ─────────────────────────────────────────────────
    ws_mm = wb.create_sheet("Dept_MinMax")
    build_dept_minmax_sheet(ws_mm, n_resources, n_depts)

    # ── Sheet 7: Working_Hours_Note ──────────────────────────────────────────
    ws_wh = wb.create_sheet("Working_Hours")
    build_working_hours_note(ws_wh)

    # ── Tab colours ──────────────────────────────────────────────────────────
    tab_colors = ["1F3864", "2E75B6", "70AD47", "ED7D31",
                  "7030A0", "C00000", "00B0F0"]
    sheets = [ws_cfg, ws_res, ws_dpt, ws_abs, ws_map, ws_mm, ws_wh]
    for ws, color in zip(sheets, tab_colors):
        ws.sheet_properties.tabColor = color

    wb.save(output_path)
    print(f"✅  Template saved → {output_path}")
    print(f"    Resources : {n_resources}")
    print(f"    Departments: {n_depts}")
    print(f"    Period    : {start_date} → {end_date}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generate the Timetable Scheduler input template.")
    parser.add_argument("--output",  default="input_template.xlsx",
                        help="Output .xlsx path (default: input_template.xlsx)")
    parser.add_argument("--resources", type=int, default=10,
                        help="Number of resources/people (default: 10)")
    parser.add_argument("--depts", type=int, default=4,
                        help="Number of departments (default: 4)")
    parser.add_argument("--start", default=None,
                        help="Plan start date YYYY-MM-DD (default: first of current month)")
    parser.add_argument("--end",   default=None,
                        help="Plan end date YYYY-MM-DD (default: last of current month)")
    args = parser.parse_args()

    start = (datetime.date.fromisoformat(args.start)
             if args.start else None)
    end   = (datetime.date.fromisoformat(args.end)
             if args.end   else None)

    generate_template(args.output, args.resources, args.depts, start, end)
